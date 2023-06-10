# -*- coding: utf-8 -*-

#https://stmtk358.hatenablog.com/entry/2021/09/13/083728
#pip install googletrans==4.0.0-rc1

import os
import sys
import openpyxl
from googletrans import Translator

args = sys.argv

lang_src="en"
lang_dest="ja"

translator = Translator()

#ファイル名
path_in=args[1]
tmpf=os.path.splitext(path_in)
path_out=tmpf[0] + '_trans' +tmpf[1]
print("input: %s\n" %(path_in))

#Excel読み込み
wb = openpyxl.load_workbook(path_in)
sheet = wb.worksheets[0]

# 1行ごとに翻訳
for row in range(1, sheet.max_row + 1):
    en = sheet.cell(row=row, column=1).value
    if en!=None:
        ja = translator.translate(text=en, src=lang_src, dest=lang_dest)
        print(" %3d: %s" % (row,en))
        print("      => %s" %(ja.text))

        sheet.cell(row=row, column=2).value = ja.text
        sheet.cell(row=row, column=2).alignment = openpyxl.styles.Alignment(wrapText=False)

#Excel出力
print("\noutput: %s" % (path_out))
wb.save(path_out)
